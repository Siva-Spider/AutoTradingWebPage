# backend/app/trading/trading_core.py

import asyncio
import pandas as pd
import ta
import datetime
import pytz
from collections import deque
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import logging
from typing import Any, Optional  # Use Any for generic broker_adapter type
import requests  # Needed for Telegram message sending

logger = logging.getLogger(__name__)


# --- Helper Functions (Broker-Agnostic or passed from broker) ---

async def send_telegram_message_async(message: str):
    """Async wrapper for sending a message to the specified Telegram channel."""
    await asyncio.to_thread(_send_telegram_message_sync, message)


def _send_telegram_message_sync(message: str):
    """Synchronous function to send a message to the specified Telegram channel."""
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN", "8102159783:AAENTxu-hXJn-uoI9TsR7tB72Cn6R1iRPX0")
    chat_id = os.getenv("TELEGRAM_CHAT_ID", "744141920")

    if not bot_token or not chat_id:
        logger.error("Error: Telegram configuration missing. Message not sent.")
        return
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    payload = {"chat_id": chat_id, "text": message, "parse_mode": "Markdown"}
    try:
        response = requests.post(url, json=payload)
        response.raise_for_status()
        logger.info("Telegram message sent.")
    except requests.exceptions.RequestException as e:
        logger.error(f"Error sending Telegram message: {e}")


def _round_to_next_interval(interval_minutes: int, now_dt: datetime.datetime) -> datetime.datetime:
    """Calculates the next interval end time."""
    base = datetime.datetime.combine(now_dt.date(), datetime.time(9, 15))  # Market open time
    elapsed = (now_dt - base).total_seconds()

    if elapsed < 0:  # Before market open, next interval is the first one
        return base.replace(microsecond=0)

    # Calculate how many full intervals have passed
    intervals_passed = int(elapsed // (interval_minutes * 60))
    # Next interval end is 'intervals_passed + 1' from base
    next_interval = base + datetime.timedelta(minutes=(intervals_passed + 1) * interval_minutes)

    return next_interval.replace(microsecond=0)


# Non-blocking indicator calculations
def calculate_supertrend(df: pd.DataFrame, period: int = 7, multiplier: int = 3) -> pd.DataFrame:
    """Calculates Supertrend indicator."""
    atr = ta.volatility.AverageTrueRange(df['high'], df['low'], df['close'], window=period).average_true_range()
    hl2 = (df['high'] + df['low']) / 2
    upperband = hl2 + (multiplier * atr)
    lowerband = hl2 - (multiplier * atr)

    supertrend_values = [True] * len(df)  # Bullish by default
    final_upperband = upperband.copy()
    final_lowerband = lowerband.copy()

    for i in range(1, len(df)):
        if df['close'].iloc[i] > final_upperband.iloc[i - 1]:
            supertrend_values[i] = True
        elif df['close'].iloc[i] < final_lowerband.iloc[i - 1]:
            supertrend_values[i] = False
        else:
            supertrend_values[i] = supertrend_values[i - 1]
            # Adjust bands if current supertrend value is same as previous
            if supertrend_values[i] and final_lowerband.iloc[i] < final_lowerband.iloc[i - 1]:
                final_lowerband.iloc[i] = final_lowerband.iloc[i - 1]
            if not supertrend_values[i] and final_upperband.iloc[i] > final_upperband.iloc[i - 1]:
                final_upperband.iloc[i] = final_upperband.iloc[i - 1]

    df['supertrend'] = [final_lowerband[i] if st else final_upperband[i] for i, st in enumerate(supertrend_values)]
    return df


async def log_order_to_excel_async(order_type: str, instrument_key: str, price: float, lots: int, strike: float,
                                   option_type: str, lot_size: int, first_expiry: datetime.date):
    """Async wrapper for logging order details to Excel."""
    await asyncio.to_thread(_log_order_to_excel_sync, order_type, instrument_key, price, lots, strike, option_type,
                            lot_size, first_expiry)


def _log_order_to_excel_sync(order_type: str, instrument_key: str, price: float, lots: int, strike: float,
                             option_type: str, lot_size: int, first_expiry: datetime.date):
    """Synchronous function for logging order details to Excel."""
    filename = "order_logs_ASMW10.xlsx"
    ist = pytz.timezone('Asia/Kolkata')
    now = datetime.datetime.now(ist)
    data = {
        'Time': [now.strftime('%Y-%m-%d %H:%M:%S')],
        'Order Type': [order_type],
        'Instrument Key': [instrument_key],
        'strike': [strike],
        'Option Type': [option_type],
        'Expiry_Date': [first_expiry.strftime('%Y-%m-%d')],
        'Lots': [lots],
        'Lot_Size': [lot_size],
        'Price': [price],
        'Live_Price': [''],
        'P&L': [''],
        'Status': ['Running']
    }
    df_new = pd.DataFrame(data)

    if os.path.exists(filename):
        with pd.ExcelWriter(filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            book = load_workbook(filename)
            sheet = book.active
            startrow = sheet.max_row
            df_new.to_excel(writer, index=False, header=False, startrow=startrow)
    else:
        df_new.to_excel(filename, index=False, engine='openpyxl')
    logger.info(f"Order logged to Excel: {order_type} {instrument_key}")


async def update_pnl_in_excel_async(broker_adapter: Any, number: int):
    """Async wrapper for updating P&L and status in Excel. Needs broker_adapter to fetch live values."""
    await asyncio.to_thread(_update_pnl_in_excel_sync, broker_adapter, number)


def _update_pnl_in_excel_sync(broker_adapter: Any, number: int):
    """Synchronous function for updating P&L and status in Excel."""
    filename = "order_logs_ASMW10.xlsx"
    if not os.path.exists(filename):
        logger.warning(f"Excel log file not found: {filename}")
        return

    df = pd.read_excel(filename, engine='openpyxl')
    wb = load_workbook(filename)
    ws = wb.active

    for idx, row in df.iterrows():
        try:
            current_status = str(row.get("Status")).strip().lower()
            instrument_key = row['Instrument Key']

            if instrument_key == "first":  # Placeholder entry
                excel_row = idx + 2
                status_col = df.columns.get_loc('Status') + 1
                ws.cell(row=excel_row, column=status_col).value = "Completed"
                continue

            if current_status == "completed":
                continue

            entry_price = float(row['Price'])
            lots = int(row['Lots'])
            lot_size = int(row['Lot_Size'])

            live_value = asyncio.run(broker_adapter.fetch_live_option_value(instrument_key))

            if live_value is None:
                logger.warning(f"Could not get live value for {instrument_key}. Skipping P&L update.")
                continue

            pnl = round(lots * lot_size * (live_value - entry_price), 2)

            excel_row = idx + 2
            pnl_col = df.columns.get_loc('P&L') + 1
            live_price_col = df.columns.get_loc('Live_Price') + 1
            status_col = df.columns.get_loc('Status') + 1

            cell_pnl = ws.cell(row=excel_row, column=pnl_col)
            cell_pnl.value = pnl
            cell_pnl.font = Font(color="FF0000" if pnl < 0 else "00B050")

            ws.cell(row=excel_row, column=live_price_col).value = live_value

            final_status = "Completed" if number == 2 else "Running"
            ws.cell(row=excel_row, column=status_col).value = final_status

        except Exception as e:
            logger.error(f"Error updating P&L for row {idx + 2}: {e}", exc_info=True)
            continue

    asyncio.run(asyncio.to_thread(wb.save, filename))


async def start_trading_session(broker_adapter: Any, index_name: str, instrument_key_index: str, interval: int,
                                lots: int):
    """
    Main asynchronous function to run the auto-trading strategy for any broker.
    This function contains the continuous market data fetching, indicator calculation,
    and signal generation loop, relying on the broker_adapter for API calls.
    """
    ist = pytz.timezone('Asia/Kolkata')
    market_open = datetime.time(9, 15)
    market_close = datetime.time(15, 31)
    today = datetime.date.today()
    start_date_for_df_filter = today - datetime.timedelta(days=25)

    logger.info(f"Initializing trading session for {index_name} with {broker_adapter.__class__.__name__}...")
    await log_order_to_excel_async("first", "first", 0, 0, 0, "first", 0, today)

    candle_buffer = deque(maxlen=2000)  # Stores candles for the main interval (e.g., 5-min)
    live_buffer = deque(maxlen=2000)  # Stores 1-minute candles for resampling

    # --- Initial Data Fetch (Historical and Intraday) ---
    logger.info("Fetching historical data once at the start...")
    historical_df = await broker_adapter.fetch_historical_data(instrument_key_index, interval)
    if historical_df.empty:
        logger.error("Failed to fetch initial historical data. Aborting trading session.")
        await send_telegram_message_async(
            f"ERROR: Trading session for {index_name} aborted. Failed to fetch historical data.")
        return

    for dt, row in historical_df.iterrows():
        dt_aware = dt if dt.tzinfo else ist.localize(dt)
        candle_buffer.append({
            'datetime': dt_aware,
            'open': row['open'],
            'high': row['high'],
            'low': row['low'],
            'close': row['close'],
        })

    now = datetime.datetime.now(ist)
    if now.time() > market_open:
        logger.info("Fetching initial intraday data up to current time...")
        intraday_df = await broker_adapter.fetch_intraday_data(instrument_key_index, interval)
        if intraday_df is not None and not intraday_df.empty:
            for dt, row in intraday_df.iterrows():
                dt_aware = dt if dt.tzinfo else ist.localize(dt)
                live_buffer.append({
                    'datetime': dt_aware,
                    'open': row['open'],
                    'high': row['high'],
                    'low': row['low'],
                    'close': row['close'],
                })
        else:
            logger.warning("⚠️ Intraday data is empty or fetch failed initially.")

    # --- Initial Indicator Calculation ---
    if candle_buffer:
        df = pd.DataFrame(list(candle_buffer))
        df['datetime'] = pd.to_datetime(df['datetime']).dt.tz_localize(None)
        df.set_index('datetime', inplace=True)
        df['adx'] = ta.trend.ADXIndicator(df['high'], df['low'], df['close'], window=14).adx()
        df['adxema'] = df['adx'].ewm(span=21, adjust=False).mean()
        df['willr'] = ta.momentum.WilliamsRIndicator(df['high'], df['low'], df['close'], lbp=14).williams_r()
        df = calculate_supertrend(df.copy(), period=7, multiplier=3)
        macd = ta.trend.MACD(df['close'], window_slow=26, window_fast=12, window_sign=9)
        df['macd'] = macd.macd()
        df['macd_signal'] = macd.macd_signal()
        df.dropna(inplace=True)
        df = df[df.index >= ist.localize(datetime.datetime.combine(start_date_for_df_filter, datetime.time.min))]

        if len(df) >= 2:  # Check length after dropping NA
            logger.info(f"Initial indicators calculated. DataFrame length: {len(df)}")
        else:
            logger.error(
                "⚠️ Not enough data for initial indicator calculation after filtering. Trading session cannot proceed.")
            await send_telegram_message_async(
                f"ERROR: Trading session for {index_name} aborted. Not enough data for indicators.")
            return
    else:
        logger.error("⚠️ Candle buffer is empty after initial fetch. Trading session cannot proceed.")
        await send_telegram_message_async(f"ERROR: Trading session for {index_name} aborted. Empty candle buffer.")
        return

    prev_time = candle_buffer[-1]['datetime'] if candle_buffer else ist.localize(
        datetime.datetime.now().replace(hour=9, minute=15, second=0, microsecond=0))
    logger.info(f"Previous Main Candle Time : {prev_time}")

    last_candle_time_1min = ist.localize(
        datetime.datetime.now().replace(second=0, microsecond=0) - datetime.timedelta(minutes=1))

    if datetime.datetime.now(ist).time() > market_open:
        logger.info("Populating 1-min live buffer with current intraday data...")
        intraday_live_1min = await broker_adapter.fetch_intraday_data(instrument_key_index, 1)  # Fetch 1-min interval
        if intraday_live_1min is not None and not intraday_live_1min.empty:
            for dt, row in intraday_live_1min.iterrows():
                dt_aware = dt if dt.tzinfo else ist.localize(dt)
                live_buffer.append({
                    'datetime': dt_aware,
                    'open': row['open'],
                    'high': row['high'],
                    'low': row['low'],
                    'close': row['close'],
                })
            if live_buffer:
                last_candle_time_1min = live_buffer[-1]['datetime']

    now_current = datetime.datetime.now(ist)
    next_interval_end_1min = _round_to_next_interval(1, now_current)
    next_resample_interval_end = _round_to_next_interval(interval, prev_time)

    logger.info(
        f"Next 1-min Candle End Time: {next_interval_end_1min}, Next {interval}-min Resample End Time: {next_resample_interval_end}")
    await send_telegram_message_async(
        f"Trading session started for {index_name} at {interval}-min interval. Waiting for market open.")

    # Wait for market open (non-blocking)
    while datetime.datetime.now(ist).time() < market_open:
        current_time_str = datetime.datetime.now(ist).strftime('%H:%M:%S')
        logger.info(f"🕘 Market hasn't opened yet. Current time: {current_time_str}")
        await asyncio.sleep(10)

    logger.info(f"✅ Market is now open! Time: {datetime.datetime.now(ist).strftime('%H:%M:%S')}")

    # --- Main Live Trading Loop ---
    while True:
        now = datetime.datetime.now(ist)
        current_time = now.time()

        if current_time >= market_close:
            logger.info("🔚 Market closed. Exiting trading loop.")
            await send_telegram_message_async(f"Trading session for {index_name} ended. Market closed.")
            await update_pnl_in_excel_async(broker_adapter, 2)
            break

        if now >= next_interval_end_1min:
            candle = await broker_adapter.fetch_ohlc_1min_data(instrument_key_index)
            if candle:
                latest_time = candle['datetime']
                if not live_buffer or latest_time > live_buffer[-1]['datetime']:
                    live_buffer.append(candle)
                    logger.info(f"New 1-min candle added to live_buffer: {latest_time}")
                else:
                    logger.debug(
                        f"Received old 1-min candle data ({latest_time} <= {live_buffer[-1]['datetime']}). Waiting...")
                next_interval_end_1min = _round_to_next_interval(1, now)
            else:
                logger.warning("⚠️ Failed to fetch new 1-min candle.")
                await asyncio.sleep(5)
                continue

        if now >= next_resample_interval_end and len(live_buffer) >= interval:
            await update_pnl_in_excel_async(broker_adapter, 1)

            df_live = pd.DataFrame(list(live_buffer)).set_index('datetime')
            df_live.index = pd.to_datetime(df_live.index)

            resampled = df_live.resample(f'{interval}min', label='left', closed='left').agg({
                'open': 'first',
                'high': 'max',
                'low': 'min',
                'close': 'last'
            }).dropna()

            if not resampled.empty:
                new_main_candles = resampled[resampled.index > prev_time]
                for timestamp, row in new_main_candles.iterrows():
                    candle_buffer.append({
                        'datetime': timestamp,
                        'open': row['open'],
                        'high': row['high'],
                        'low': row['low'],
                        'close': row['close']
                    })
                if not new_main_candles.empty:
                    prev_time = new_main_candles.index[-1]
                    logger.info(f"New {interval}-min candle added to candle_buffer: {prev_time}")
                else:
                    logger.debug(f"No new {interval}-min candles after resampling.")
            else:
                logger.warning("Resampled DataFrame is empty.")

            if len(candle_buffer) >= 2:
                df = pd.DataFrame(list(candle_buffer)).set_index('datetime')
                df['adx'] = ta.trend.ADXIndicator(df['high'], df['low'], df['close'], window=14).adx()
                df['adxema'] = df['adx'].ewm(span=21, adjust=False).mean()
                df['willr'] = ta.momentum.WilliamsRIndicator(df['high'], df['low'], df['close'], lbp=14).williams_r()
                df = calculate_supertrend(df.copy(), period=7, multiplier=3)
                macd = ta.trend.MACD(df['close'], window_slow=26, window_fast=12, window_sign=9)
                df['macd'] = macd.macd()
                df['macd_signal'] = macd.macd_signal()
                df.dropna(inplace=True)
                df_filtered = df[
                    df.index >= ist.localize(datetime.datetime.combine(start_date_for_df_filter, datetime.time.min))]

                if len(df_filtered) >= 2:
                    latest_candle = df_filtered.iloc[-1]
                    latest_willr = latest_candle['willr']
                    latest_macd = latest_candle['macd']
                    latest_macd_signal = latest_candle['macd_signal']
                    latest_adx = latest_candle['adx']
                    latest_adxema = latest_candle['adxema']
                    latest_supertrend = latest_candle['supertrend']
                    close_price = latest_candle['close']

                    logger.info(
                        f"Indicators at {latest_candle.name}: ADX={latest_adx:.2f}, WILLR={latest_willr:.2f}, MACD={latest_macd:.2f}, Supertrend={latest_supertrend:.2f}")

                    # ✅ Check for trade signal conditions (common logic for all brokers)
                    positions = await broker_adapter.fetch_positions()

                    buy_signal = (latest_adx > latest_adxema and latest_willr > -30 and
                                  latest_supertrend < close_price and latest_macd > latest_macd_signal)
                    sell_signal = (latest_adx > latest_adxema and latest_willr < -70 and
                                   latest_supertrend > close_price and latest_macd < latest_macd_signal)

                    if buy_signal:
                        logger.info("🔼 BUY SIGNAL GENERATED")
                        await send_telegram_message_async(f"BUY SIGNAL GENERATED for {index_name}.")
                        has_ce_position = any(
                            pos['quantity'] > 0 and pos['tradingsymbol'].endswith("CE") for pos in positions)
                        if not has_ce_position:
                            option_details = await broker_adapter.get_option_instrument_key(close_price, "CE",
                                                                                            index_name, interval)
                            if option_details:
                                await broker_adapter.place_order(
                                    option_details["instrument_key"],
                                    lots * option_details["lot_size"],
                                    "BUY",
                                    close_price
                                )
                                await log_order_to_excel_async("BUY", option_details["instrument_key"], close_price,
                                                               lots, option_details["strike"], "CE",
                                                               option_details["lot_size"],
                                                               option_details["expiry_date"])
                            else:
                                logger.error("Failed to get CE option details for order placement.")
                        else:
                            logger.info(f"Existing CE position found for {index_name}. Skipping new BUY order.")
                    elif sell_signal:
                        logger.info("🔽 SELL SIGNAL GENERATED")
                        await send_telegram_message_async(f"SELL SIGNAL GENERATED for {index_name}.")
                        has_pe_position = any(
                            pos['quantity'] > 0 and pos['tradingsymbol'].endswith("PE") for pos in positions)
                        if not has_pe_position:
                            option_details = await broker_adapter.get_option_instrument_key(close_price, "PE",
                                                                                            index_name, interval)
                            if option_details:
                                await broker_adapter.place_order(
                                    option_details["instrument_key"],
                                    lots * option_details["lot_size"],
                                    "BUY",
                                    close_price
                                )
                                await log_order_to_excel_async("BUY", option_details["instrument_key"], close_price,
                                                               lots, option_details["strike"], "PE",
                                                               option_details["lot_size"],
                                                               option_details["expiry_date"])
                            else:
                                logger.error("Failed to get PE option details for order placement.")
                        else:
                            logger.info(f"Existing PE position found for {index_name}. Skipping new SELL order.")
                    else:
                        logger.info("⏸️ NO TRADE SIGNAL GENERATED")

                    # ✅ Check for exit conditions (common logic for all brokers)
                    for pos in positions:
                        quantity = pos['quantity']
                        if quantity > 0:
                            instrument_token = pos['instrument_token']
                            tradingsymbol = pos['tradingsymbol']
                            option_type = tradingsymbol[-2:]

                            exit_ce = (option_type == "CE" and (
                                    (latest_willr < -70 and latest_supertrend > close_price) or
                                    (latest_willr < -70 and latest_macd < latest_macd_signal) or
                                    (latest_supertrend > close_price and latest_macd < latest_macd_signal)
                            ))
                            exit_pe = (option_type == "PE" and (
                                    (latest_willr > -30 and latest_supertrend < close_price) or
                                    (latest_willr > -30 and latest_macd > latest_macd_signal) or
                                    (latest_supertrend < close_price and latest_macd < latest_macd_signal)
                            ))

                            if exit_ce or exit_pe:
                                logger.info(f"Exit conditions triggered. Closing position {tradingsymbol}")
                                await send_telegram_message_async(
                                    f"Exit conditions triggered. Closing position {tradingsymbol} for {index_name}.")
                                await broker_adapter.place_order(instrument_token, quantity, "SELL", 0)
                                await update_pnl_in_excel_async(broker_adapter, 2)

            next_resample_interval_end = _round_to_next_interval(interval, now)
            logger.info(f"Next {interval}-min Resample End Time updated to: {next_resample_interval_end}")

        await asyncio.sleep(5)  # Poll every 5 seconds to check for new data or market close

